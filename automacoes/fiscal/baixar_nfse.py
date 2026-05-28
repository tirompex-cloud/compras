"""
Automação de Download de NFS-e + Geração de Planilha
=====================================================
Pré-requisitos (rodar UMA vez no CMD):
  python -m pip install playwright openpyxl
  python -m playwright install chrome

Como usar todo mês:
  python baixar_nfse.py
"""

import os
import sys
import glob
import time
from datetime import datetime
from urllib.parse import quote
from xml.etree import ElementTree as ET

try:
    from playwright.sync_api import sync_playwright
except ImportError:
    print("❌ Playwright não instalado. Rode:")
    print("   python -m pip install playwright")
    print("   python -m playwright install chrome")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("❌ openpyxl não instalado. Rode:  python -m pip install openpyxl")
    sys.exit(1)


URL_LOGIN  = "https://www.nfse.gov.br/EmissorNacional/Login"
PASTA_BASE = os.path.join(os.path.expanduser("~"), "Desktop", "NFS_Downloads")
TIMEOUT_DL = 60_000   # 60s para downloads
PAUSA      = 1.2      # segundos entre cliques


def url_notas(data_inicio, data_fim):
    """
    Monta a URL com o filtro de datas já aplicado.
    data_inicio / data_fim no formato DD/MM/AAAA
    """
    di = quote(data_inicio, safe="")   # ex: 01%2F03%2F2026
    df = quote(data_fim,    safe="")
    return (
        f"https://www.nfse.gov.br/EmissorNacional/Notas/Recebidas"
        f"?executar=1&busca=&datainicio={di}&datafim={df}"
    )


# ════════════════════════════════════════════════════════════════════════════
#  EXTRAÇÃO DE DADOS DO XML
# ════════════════════════════════════════════════════════════════════════════
def find_text(root, *tags):
    for tag in tags:
        for elem in root.iter():
            local = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
            if local == tag and elem.text:
                return elem.text.strip()
    return None


def extrair_dados_xml(xml_path):
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()
    except Exception:
        return None

    # Valor — padrão nacional usa ponto como decimal (ex: 350.00)
    valor_str = find_text(root,
        "vLiq",           # valor líquido — tag principal do padrão nacional
        "vServ",          # valor do serviço dentro de vServPrest
        "ValorServicos", "ValorTotal", "ValorLiquidoNfse", "ValorNfse",
    )
    try:
        if valor_str:
            # XML nacional usa ponto como decimal — não remover o ponto!
            valor = float(valor_str.replace(",", "."))
        else:
            valor = None
    except Exception:
        valor = None

    # Data — padrão nacional usa dhEmi com timezone (ex: 2026-03-24T13:35:55-03:00)
    data_str = find_text(root, "dhEmi", "dCompet", "DataEmissao", "DataEmissaoNfse")
    emissao = None
    if data_str:
        # Remove timezone se existir: 2026-03-24T13:35:55-03:00 → 2026-03-24T13:35:55
        data_str = data_str[:19]
        for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                emissao = datetime.strptime(data_str, fmt)
                break
            except ValueError:
                continue

    # Empresa — o emitente da nota (prestador do serviço)
    empresa = find_text(root, "xNome", "RazaoSocial", "NomeFantasia")

    # Número da NF — prefere o número curto (nNFSe) ao ID longo
    nf  = find_text(root, "nNFSe", "Numero", "NumeroNfse")
    rps = find_text(root, "nDPS", "NumeroRps", "NumRps")

    return {
        "nf":      nf,
        "rps":     rps if not nf else None,
        "empresa": empresa,
        "valor":   valor,
        "emissao": emissao,
    }


# ════════════════════════════════════════════════════════════════════════════
#  GERAÇÃO DA PLANILHA
# ════════════════════════════════════════════════════════════════════════════
def gerar_planilha(pasta, mes_ano):
    xmls = glob.glob(os.path.join(pasta, "**", "*.xml"), recursive=True)
    if not xmls:
        print("⚠  Nenhum XML encontrado para gerar planilha.")
        return None

    registros = [extrair_dados_xml(x) for x in xmls]
    registros = [r for r in registros if r]
    registros.sort(key=lambda r: (
        r.get("emissao") or datetime(1900, 1, 1),
        int(r.get("nf") or r.get("rps") or 0)
    ))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planilha1"

    hfont  = Font(name="Arial", bold=True, size=11)
    dfont  = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")
    right  = Alignment(horizontal="right",  vertical="center")
    thin   = Side(style="thin", color="BFBFBF")
    brd    = Border(left=thin, right=thin, top=thin, bottom=thin)
    hfill  = PatternFill("solid", fgColor="D9E1F2")
    afill  = PatternFill("solid", fgColor="F2F2F2")

    for col, h in enumerate(["", "NF", "FATURA", "EMPRESA", "VALOR", "EMISSÃO"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hfont; c.fill = hfill; c.border = brd; c.alignment = center

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14

    for i, reg in enumerate(registros, 2):
        fill = afill if i % 2 == 0 else PatternFill()
        nf = reg.get("nf"); rps = reg.get("rps")

        def cel(col, val, aln, _i=i, _f=fill):
            c = ws.cell(row=_i, column=col, value=val)
            c.font = dfont; c.fill = _f; c.border = brd; c.alignment = aln
            return c

        ws.cell(row=i, column=1).fill = fill
        # NF como texto para evitar notação científica
        c_nf = ws.cell(row=_i if False else i, column=2)
        c_nf.value = str(int(float(nf))) if nf and str(nf).replace('.','').isdigit() else nf
        c_nf.font = dfont; c_nf.fill = fill; c_nf.border = brd; c_nf.alignment = center
        c_nf.number_format = '@'  # formato texto

        c_rps = ws.cell(row=i, column=3)
        c_rps.value = str(int(float(rps))) if (rps and not nf and str(rps).replace('.','').isdigit()) else (rps if not nf else None)
        c_rps.font = dfont; c_rps.fill = fill; c_rps.border = brd; c_rps.alignment = center
        c_rps.number_format = '@'
        cel(4, (reg.get("empresa") or "").upper(),       left)
        c = cel(5, reg.get("valor"),                     right);  c.number_format = "#,##0.00"
        c = cel(6, reg.get("emissao"),                   center); c.number_format = "DD/MM/YYYY"

    last = len(registros) + 2
    yfill = PatternFill("solid", fgColor="FFD966")
    for col in range(1, 7):
        ws.cell(row=last, column=col).fill = yfill
        ws.cell(row=last, column=col).border = brd
    lbl = ws.cell(row=last, column=4, value="TOTAL")
    lbl.font = Font(name="Arial", bold=True, size=11); lbl.alignment = right
    tot = ws.cell(row=last, column=5, value=f"=SUM(E2:E{last-1})")
    tot.number_format = "#,##0.00"
    tot.font = Font(name="Arial", bold=True, size=11); tot.alignment = right

    nome    = f"Planilha_NFS_-_{mes_ano}.xlsx"
    caminho = os.path.join(pasta, nome)
    wb.save(caminho)
    total_val = sum(r["valor"] or 0 for r in registros)
    print(f"\n✅ Planilha gerada: {caminho}")
    print(f"   📊 {len(registros)} nota(s) | 💰 Total: R$ {total_val:,.2f}")
    return caminho


# ════════════════════════════════════════════════════════════════════════════
#  EXTRAÇÃO DE LINKS DO DOM
# ════════════════════════════════════════════════════════════════════════════
def extrair_links_download(page):
    """
    Extrai todos os links de download diretamente do DOM (sem abrir menus).
    O portal esconde os itens do menu em div.menu-content — eles já estão no HTML.
    Retorna lista de dicts com as URLs de cada nota.
    """
    links = page.evaluate("""
        () => {
            const notas = [];
            const rows = document.querySelectorAll('tbody tr');
            rows.forEach((row, idx) => {
                const menu = row.querySelector('.menu-content, .list-group');
                if (!menu) return;
                const anchors = menu.querySelectorAll('a');
                const nota = { linha: idx + 1, links: [] };
                anchors.forEach(a => {
                    nota.links.push({ href: a.href, texto: a.textContent.trim() });
                });
                notas.push(nota);
            });
            return notas;
        }
    """)
    return links


# ════════════════════════════════════════════════════════════════════════════
#  DOWNLOAD DE UMA NOTA (PDF + XML via links diretos)
# ════════════════════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════════════════════
#  DOWNLOAD DE UMA NOTA (PDF + XML via links diretos)
# ════════════════════════════════════════════════════════════════════════════
def baixar_arquivo(page, context, href, pasta_download, nome_base, ext):
    """Faz download abrindo uma nova aba pelo Playwright."""
    try:
        nova_aba = context.new_page()
        with nova_aba.expect_download(timeout=TIMEOUT_DL) as dl_info:
            try:
                nova_aba.goto(href, wait_until="domcontentloaded", timeout=30_000)
            except Exception:
                pass  # Download já iniciou, ignora erro de navegação
        dl = dl_info.value
        nome = dl.suggested_filename or f"{nome_base}.{ext}"
        dl.save_as(os.path.join(pasta_download, nome))
        nova_aba.close()
        icone = "📄" if ext == "pdf" else "📋"
        print(f"    {icone} {ext.upper()}: {nome}")
        return True
    except Exception as e:
        print(f"    ⚠ Erro ao baixar {ext.upper()}: {e}")
        try:
            nova_aba.close()
        except Exception:
            pass
        return False


def baixar_nota_por_links(page, context, nota_info, pasta_download, numero_nota):
    baixados = 0
    nome_base = f"NF_{numero_nota:04d}"

    for item in nota_info.get("links", []):
        href  = item.get("href", "")
        texto = item.get("texto", "")

        if not href or href.startswith("javascript"):
            continue
        if any(x in texto for x in ["Visualizar", "Confirmar", "Rejeitar"]):
            continue

        if "DANFS" in texto or "DANFS" in href:
            ext = "pdf"
        elif "XML" in texto or "NFSe" in href:
            ext = "xml"
        else:
            continue

        # Pula se já existe arquivo com esse nome base
        ja_existe = any(
            f.startswith(nome_base) and f.endswith(f".{ext}")
            for f in os.listdir(pasta_download)
        )
        if ja_existe:
            print(f"    ⏭ {ext.upper()} já existe, pulando")
            continue

        ok = baixar_arquivo(page, context, href, pasta_download, nome_base, ext)
        if ok:
            baixados += 1
        time.sleep(0.8)

    return baixados


# ════════════════════════════════════════════════════════════════════════════
#  NAVEGAÇÃO ENTRE PÁGINAS
# ════════════════════════════════════════════════════════════════════════════
def contar_linhas(page):
    try:
        return page.locator("tbody tr").count()
    except Exception:
        return 0


def proxima_pagina(page):
    """Clica no penúltimo botão da paginação (sempre é o ›). Retorna False na última página."""

    page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(0.5)

    # Captura link da 1ª nota antes de clicar
    antes = page.evaluate("""
        () => {
            const a = document.querySelector('tbody tr .menu-content a');
            return a ? a.href : document.querySelector('tbody tr')?.innerHTML?.substring(0, 100) || '';
        }
    """)

    # Log dos itens de paginação para debug
    debug = page.evaluate("""
        () => {
            const items = Array.from(document.querySelectorAll('.pagination li'));
            return items.map((li, i) => {
                const a = li.querySelector('a');
                return i + ':' + (a ? JSON.stringify(a.textContent.trim()) : 'sem-a') + ':' + li.className;
            }).join(' | ');
        }
    """)
    print(f"    [pag debug] {debug}")

    # Clica no penúltimo li (índice -2, que é o ›)
    clicou = page.evaluate("""
        () => {
            const items = Array.from(document.querySelectorAll('.pagination li'));
            if (items.length < 2) return 'sem-itens';
            const penultimo = items[items.length - 2];
            const a = penultimo.querySelector('a');
            if (!a) return 'sem-link';
            // Se estiver desabilitado E o texto for ›/>, pode ser última página
            const txt = a.textContent.trim();
            a.click();
            return 'clicou:' + txt + ':disabled=' + penultimo.classList.contains('disabled');
        }
    """)
    print(f"    [pag clique] {clicou}")

    if "sem-itens" in clicou or "sem-link" in clicou:
        return False

    time.sleep(3)

    # Verifica se o conteúdo mudou
    depois = page.evaluate("""
        () => {
            const a = document.querySelector('tbody tr .menu-content a');
            return a ? a.href : document.querySelector('tbody tr')?.innerHTML?.substring(0, 100) || '';
        }
    """)

    mudou = depois != antes and depois != ''
    print(f"    [pag mudou] {mudou}")
    return mudou


def total_registros_portal(page):
    """Lê o 'Total de X registros' que o portal mostra."""
    try:
        txt = page.evaluate("""
            () => {
                const els = document.querySelectorAll('*');
                for (const el of els) {
                    if (el.children.length === 0 && /total de \\d+/i.test(el.textContent)) {
                        return el.textContent.trim();
                    }
                }
                return '';
            }
        """)
        import re
        m = re.search(r'\d+', txt)
        return int(m.group()) if m else None
    except Exception:
        return None


# ════════════════════════════════════════════════════════════════════════════
#  MAIN
# ════════════════════════════════════════════════════════════════════════════
def main():
    print("=" * 60)
    print("  Automação NFS-e — Download + Planilha")
    print("=" * 60)

    # ── Datas ────────────────────────────────────────────────────────────────
    print("\n📅 Período das notas fiscais:")

    while True:
        data_inicio = input("  Data inicial (DD/MM/AAAA, ex: 01/03/2026): ").strip()
        try:
            datetime.strptime(data_inicio, "%d/%m/%Y")
            break
        except ValueError:
            print("  ❌ Data inválida. Use o formato DD/MM/AAAA.")

    while True:
        data_fim = input("  Data final   (DD/MM/AAAA, ex: 31/03/2026): ").strip()
        try:
            datetime.strptime(data_fim, "%d/%m/%Y")
            break
        except ValueError:
            print("  ❌ Data inválida. Use o formato DD/MM/AAAA.")

    # Valida e extrai mês/ano
    try:
        dt_fim  = datetime.strptime(data_fim, "%d/%m/%Y")
        mes_ano = dt_fim.strftime("%m_%Y")
    except ValueError:
        mes_ano = datetime.now().strftime("%m_%Y")

    pasta_download = os.path.join(PASTA_BASE, f"NFS_{mes_ano}")
    os.makedirs(pasta_download, exist_ok=True)
    print(f"\n📂 Arquivos serão salvos em:\n   {pasta_download}")

    # URL já com filtro de datas aplicado
    url_lista = url_notas(data_inicio, data_fim)
    print(f"\n🔗 URL que será usada após login:\n   {url_lista}")

    with sync_playwright() as p:
        print("\n🌐 Abrindo Chrome...")
        browser = p.chromium.launch(
            channel="chrome",       # Chrome instalado no PC = acesso ao certificado
            headless=False,
            args=["--start-maximized"],
        )
        context = browser.new_context(accept_downloads=True, viewport=None)
        page    = context.new_page()

        # ── Login com certificado ─────────────────────────────────────────────
        page.goto(URL_LOGIN, timeout=60_000)

        print("\n" + "─" * 60)
        print("  Faça o login com seu certificado digital no Chrome.")
        print("  Quando entrar no portal, volte aqui e pressione Enter.")
        print("─" * 60)
        input("  ▶ Enter após fazer login... ")

        # ── Navega para lista de notas com datas na URL ───────────────────────
        print(f"\n🔍 Carregando lista de notas ({data_inicio} a {data_fim})...")
        try:
            page.goto(url_lista, timeout=30_000, wait_until="domcontentloaded")
        except Exception:
            pass
        time.sleep(3)

        # Se não carregou notas, tenta preencher o formulário via JavaScript
        if contar_linhas(page) == 0:
            print("  Preenchendo formulário via JavaScript...")
            page.evaluate(f"""
                () => {{
                    // Preenche campos de data por diferentes atributos
                    const campos = document.querySelectorAll('input[type="text"], input[type="date"]');
                    const setVal = (el, val) => {{
                        const nativeInput = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value');
                        nativeInput.set.call(el, val);
                        el.dispatchEvent(new Event('input', {{ bubbles: true }}));
                        el.dispatchEvent(new Event('change', {{ bubbles: true }}));
                    }};
                    campos.forEach((el, i) => {{
                        const label = (el.name + el.id + el.placeholder + el.className).toLowerCase();
                        if (label.includes('inic') || label.includes('de') || i === 0) setVal(el, '{data_inicio}');
                        if (label.includes('fim') || label.includes('ate') || label.includes('final') || i === 1) setVal(el, '{data_fim}');
                    }});
                    // Clica no botão Filtrar
                    const btns = document.querySelectorAll('button, input[type=submit]');
                    btns.forEach(b => {{ if ((b.textContent || b.value || '').includes('Filtrar')) b.click(); }});
                }}
            """)
            time.sleep(4)

        # Se ainda não carregou, pede para o usuário filtrar manualmente
        if contar_linhas(page) == 0:
            print("\n⚠ Script não conseguiu preencher as datas automaticamente.")
            print("  Por favor, preencha as datas no Chrome e clique em Filtrar.")
            input("  ▶ Pressione Enter após as notas aparecerem na tela... ")

        total_na_pagina = contar_linhas(page)
        if total_na_pagina == 0:
            print("\n❌ Nenhuma nota encontrada. Verifique as datas e tente novamente.")
            browser.close()
            return
        print(f"  ✔ {total_na_pagina} nota(s) carregada(s)")

        # ── Scroll para carregar todas as notas (portais com lazy loading) ──────
        print("  🔄 Verificando se há mais notas para carregar...")
        ultima_contagem = 0
        for _ in range(20):  # máx 20 scrolls
            contagem = contar_linhas(page)
            if contagem == ultima_contagem:
                break
            ultima_contagem = contagem
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            time.sleep(1.5)
        print(f"  ✔ Total de notas na página: {contar_linhas(page)}")

        # ── FASE 1: Coletar links de todas as páginas ─────────────────────────
        print("\n📋 Coletando links de todas as páginas...")

        total_esperado = total_registros_portal(page)
        if total_esperado:
            print(f"  Portal indica: {total_esperado} notas no total")

        todas_notas = []
        pagina_atual = 1

        while True:
            time.sleep(1.5)
            notas_pagina = extrair_links_download(page)
            todas_notas.extend(notas_pagina)
            print(f"  Página {pagina_atual}: {len(notas_pagina)} nota(s) — acumulado: {len(todas_notas)}")

            # Verifica se já coletamos tudo
            if total_esperado and len(todas_notas) >= total_esperado:
                print(f"  ✔ Total esperado atingido ({total_esperado} notas)")
                break

            if proxima_pagina(page):
                pagina_atual += 1
            else:
                print(f"  ✔ Última página ({pagina_atual}), total: {len(todas_notas)} nota(s)")
                break

        # ── FASE 2: Baixar todos os arquivos ──────────────────────────────────
        print("⬇  Iniciando downloads...\n")
        total_notas = 0
        total_arqs  = 0

        for nota_info in todas_notas:
            total_notas += 1
            print(f"  📝 Nota #{total_notas}/{len(todas_notas)}")
            arqs = baixar_nota_por_links(page, context, nota_info, pasta_download, total_notas)
            total_arqs += arqs
            time.sleep(0.3)

        print(f"\n{'='*60}")
        print(f"  ✅ {total_notas} nota(s) | {total_arqs} arquivo(s) baixado(s)")
        print(f"  📁 {pasta_download}")
        browser.close()

    # ── Gerar planilha ────────────────────────────────────────────────────────
    print(f"\n📊 Gerando planilha Excel...")
    caminho = gerar_planilha(pasta_download, mes_ano)

    print(f"\n🎉 Tudo pronto!")
    if caminho:
        print(f"   📊 {caminho}")

    input("\nPressione Enter para fechar...")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        import traceback
        print("\n" + "=" * 60)
        print("  ❌ ERRO INESPERADO:")
        print("=" * 60)
        traceback.print_exc()
        print("=" * 60)
        input("\nPressione Enter para fechar...")